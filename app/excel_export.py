# app/excel_export.py
# Funcionalidade de exportação para Excel - Implementação futura

import os
from pathlib import Path
from datetime import datetime

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_DISPONIVEL = True
except ImportError:
    pd = None
    openpyxl = None
    EXCEL_DISPONIVEL = False

def verificar_dependencias_excel():
    """
    Verifica se as dependências para Excel estão instaladas.
    """
    return EXCEL_DISPONIVEL

def transferir_planilha(resumo_dados, file_path=None, flet_page=None, save_excel_dialog_func=None):
    """
    Abre janela para salvar planilha Excel com os dados da análise.
    """
    # Verificar se as bibliotecas estão disponíveis
    if not verificar_dependencias_excel():
        if flet_page and hasattr(flet_page, 'add') and callable(flet_page.add):
            import flet as ft
            flet_page.add(ft.SnackBar(ft.Text("Bibliotecas necessárias não instaladas. Execute: pip install pandas openpyxl"), bgcolor="#ff9800"))
            if hasattr(flet_page, 'update') and callable(flet_page.update):
                flet_page.update()
        return
        
    if not resumo_dados:
        if flet_page and hasattr(flet_page, 'add') and callable(flet_page.add):
            import flet as ft
            flet_page.add(ft.SnackBar(ft.Text("Nenhum dado para exportar"), bgcolor="#d32f2f"))
            if hasattr(flet_page, 'update') and callable(flet_page.update):
                flet_page.update()
        return
    
    if save_excel_dialog_func and callable(save_excel_dialog_func):
        save_excel_dialog_func()
    else:
        if flet_page and hasattr(flet_page, 'add') and callable(flet_page.add):
            import flet as ft
            flet_page.add(ft.SnackBar(ft.Text("Função de salvar não está disponível"), bgcolor="#d32f2f"))
            if hasattr(flet_page, 'update') and callable(flet_page.update):
                flet_page.update()

def salvar_excel_em_local(caminho_arquivo, resumo_dados, flet_page=None):
    """
    Cria uma planilha Excel (.xlsx) profissional com os dados da análise no local especificado.
    """
    try:
        if not resumo_dados:
            return

        # Preparar dados para a planilha
        data_atual = datetime.now().strftime("%d/%m/%Y %H:%M")
        
        # Função para limpar valores monetários para float
        def extrair_valor_numerico(valor):
            if isinstance(valor, str) and valor.startswith("R$"):
                return float(valor.replace("R$", "").replace(".", "").replace(",", ".").strip())
            return 0.0

        # Converter quantidade de páginas para int
        def converter_qtde_paginas(valor):
            if isinstance(valor, str) and valor.isdigit():
                return int(valor)
            return 0

        # Dados principais seguindo a formatação especificada
        dados_analise = {
            'Data da Análise': [data_atual],
            'Intervalo de Datas': [resumo_dados['intervalo_datas']],
            'Qtde Páginas': [converter_qtde_paginas(resumo_dados['qtd_pgs'])],
            'Valor Demonstrativo': [extrair_valor_numerico(resumo_dados['valor_demonstrativo'])],
            'Valor Furnapen': [extrair_valor_numerico(resumo_dados['valor_funarpen'])],
            'Valor ISSQN': [extrair_valor_numerico(resumo_dados['valor_issqn'])],
            'Valor Líquido Total': [extrair_valor_numerico(resumo_dados['valor_liquido'])]
        }

        # Criar DataFrame principal
        df_analise = pd.DataFrame(dados_analise)
        
        # Salvar no local especificado pelo usuário
        with pd.ExcelWriter(caminho_arquivo, engine='openpyxl') as writer:
            # Aba principal com dados da análise
            df_analise.to_excel(writer, sheet_name='Dados da Análise', index=False)
            
            # Formatação da aba principal
            workbook = writer.book
            worksheet_analise = writer.sheets['Dados da Análise']
            
            # Configurar larguras das colunas
            worksheet_analise.column_dimensions['A'].width = 18  # Data da Análise
            worksheet_analise.column_dimensions['B'].width = 25  # Intervalo de Datas
            worksheet_analise.column_dimensions['C'].width = 15  # Qtde Páginas
            worksheet_analise.column_dimensions['D'].width = 20  # Valor Demonstrativo
            worksheet_analise.column_dimensions['E'].width = 18  # Valor Furnapen
            worksheet_analise.column_dimensions['F'].width = 15  # Valor ISSQN
            worksheet_analise.column_dimensions['G'].width = 20  # Valor Líquido Total
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="008EBC", end_color="008EBC", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")
            right_alignment = Alignment(horizontal="right", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'), 
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Aplicar formatação aos cabeçalhos
            for cell in worksheet_analise[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
            
            # Aplicar formatação aos dados
            for row in worksheet_analise.iter_rows(min_row=2, max_row=len(df_analise)+1):
                for cell in row:
                    cell.border = border
                    # Colunas numéricas (C, D, E, F, G) alinhadas à direita
                    if cell.column in [3, 4, 5, 6, 7]:
                        cell.alignment = right_alignment
                    else:
                        cell.alignment = center_alignment
        
        if flet_page and hasattr(flet_page, 'add') and callable(flet_page.add):
            import flet as ft
            flet_page.add(ft.SnackBar(ft.Text(f"✓ Planilha Excel salva: {Path(caminho_arquivo).name}"), bgcolor="#4caf50"))
            if hasattr(flet_page, 'update') and callable(flet_page.update):
                flet_page.update()
                
    except Exception as e:
        if flet_page and hasattr(flet_page, 'add') and callable(flet_page.add):
            import flet as ft
            flet_page.add(ft.SnackBar(ft.Text(f"Erro ao salvar planilha: {str(e)}"), bgcolor="#d32f2f"))
            if hasattr(flet_page, 'update') and callable(flet_page.update):
                flet_page.update()

# Exemplo de como usar as funções quando for implementar:
"""
from app.excel_export import transferir_planilha, salvar_excel_em_local

# Para usar na interface:
def botao_excel_click(e=None):
    transferir_planilha(
        resumo_dados=estado["resumo"], 
        file_path=estado["file_path"],
        flet_page=flet_page,
        save_excel_dialog_func=save_excel_dialog_func
    )

# Para usar na callback do dialog:
def save_excel_result(e: ft.FilePickerResultEvent):
    if e.path:
        salvar_excel_em_local(e.path, estado["resumo"], flet_page)
""" 